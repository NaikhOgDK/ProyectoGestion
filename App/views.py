from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib import messages
from .forms import *
from .models import *
import pandas as pd
from django.http import HttpResponse
from django.db.models import Count, Q, OuterRef, Subquery, Max
from django.core.paginator import Paginator,  EmptyPage, PageNotAnInteger
from django.views.generic import ListView
from django.utils import timezone
import requests
from decouple import config
from django.http import JsonResponse
from cryptography.fernet import Fernet
import json
from datetime import datetime, timedelta, date
from dateutil import parser
import pytz  # Para manejar zonas horarias
from .utils import *
from .decorators import role_required
from django.views.generic import ListView
from django.template.loader import render_to_string
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from django.urls import reverse
import boto3
from django.conf import settings
import uuid
from django.utils.timezone import now

def subir_a_s3(archivo, carpeta="licencias"):
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Generar un nombre único para el archivo
        extension = archivo.name.split('.')[-1]  # Obtener la extensión
        nombre_archivo = f"{carpeta}/{uuid.uuid4()}.{extension}"  # Carpeta + UUID + extensión

        # Subir archivo a S3
        s3_client.upload_fileobj(archivo, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)

        print(f"Archivo subido correctamente: {nombre_archivo}")
        return nombre_archivo  # Retorna la ruta del archivo en S3
    except Exception as e:
        print(f"Error al subir archivo a S3: {e}")
        return None

@role_required(['Administrador'])
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, "Registration successful. You can now log in.")
            return redirect('home')
    else:
        form = UserRegisterForm()
    return render(request, 'acceso/register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                response = HttpResponse()
                # Set token in cookies
                encrypted_token = encrypt(user)
                response.set_cookie('auth_token', encrypted_token.decode(), httponly=True, secure=True)

                # Redirect based on role
                if user.role.name == 'Administrador':
                    response['Location'] = 'home'
                elif user.role.name == 'Visualizador':
                    response['Location'] = 'home'
                elif user.role.name == 'Empresa':
                    response['Location'] = 'homeEmpresa'
                elif user.role.name == 'Taller':
                    response['Location'] = 'homeTallerUsuario'
                elif user.role.name == 'AC Comercial':
                    response['Location'] = 'homeACComercial'
                else:
                    messages.error(request, "Role not defined for this user.")
                    return render(request, 'acceso/login.html', {'form': form})

                response.status_code = 302
                return response
            else:
                messages.error(request, "Invalid username or password.")
    else:
        form = UserLoginForm()
    return render(request, 'acceso/login.html', {'form': form})

def logout_user(request):
    logout(request)
    return render(request, "acceso/logout.html")

#Carga Informacion
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from App.models import (
    Marca, TipoVehiculo, Carroceria, Canal, Zona, Seccion, ubicacion_fisica, 
    Propietario, Operacion, Group, Microempresa, Vehiculo
)

@role_required(['Administrador'])
def cargar_datos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']

        try:
            # Leer solo la hoja llamada 'Base'
            df_base = pd.read_excel(archivo, sheet_name='Plantilla')

            # Contador para llevar registro de los registros nuevos
            registros_creados = 0

            # Iterar sobre las filas del DataFrame
            for _, row in df_base.iterrows():
                # Verificar si ya existe un registro con la misma patente
                if not Vehiculo.objects.filter(patente=row['Patente']).exists():
                    try:
                        # Crear o buscar objetos relacionados
                        marca, _ = Marca.objects.get_or_create(nombre=row['Marca'])
                        tipo_vehiculo, _ = TipoVehiculo.objects.get_or_create(tipo=row['Tipo Vehiculo'])
                        carroceria, _ = Carroceria.objects.get_or_create(nombre=row['Tipo carrocería'])
                        canal, _ = Canal.objects.get_or_create(canal=row['Canal'])
                        zona, _ = Zona.objects.get_or_create(zona=row['Zona'])
                        seccion, _ = Seccion.objects.get_or_create(nombre=row['Sección'])
                        ubicacion, _ = ubicacion_fisica.objects.get_or_create(nombre=row['Ubicación fisica'])
                        propietario, _ = Propietario.objects.get_or_create(tipo_propietario=row['Propietario'])
                        operacion, _ = Operacion.objects.get_or_create(operacion=row['Operación'])
                        empresa, _ = Group.objects.get_or_create(name=row['Empresa'])
                        tipo, _ = Tipo.objects.get_or_create(nombre=row['Tipo'])
                        
                        # Si es una microempresa, crearla o buscarla
                        es_microempresa = pd.notna(row["Transportista"]) and row["Transportista"] != "N/A"
                        microempresa = None
                        if es_microempresa:
                            microempresa, _ = Microempresa.objects.get_or_create(nombre=row["Transportista"])

                        # Crear el vehículo con todas las relaciones
                        Vehiculo.objects.create(
                            patente=row['Patente'],
                            marca=marca,
                            modelo=row['Modelo'],
                            ano=row['Año'],
                            num_motor=row['N° de motor'],
                            num_chasis=row['N° Chasis'],
                            num_pallets=row['N° de pallets'],
                            tipo_vehiculo=tipo_vehiculo,
                            carroceria=carroceria,
                            canal=canal,
                            zona=zona,
                            seccion=seccion,
                            ubicacion_fisica=ubicacion,
                            propietario=propietario,
                            operacion=operacion,
                            empresa=empresa,
                            es_microempresa=es_microempresa,
                            microempresa=microempresa,
                            tipo=tipo
                        )
                        registros_creados += 1
                    except Exception as e:
                        messages.error(request, f"Error al crear vehículo con patente {row['Patente']}: {e}")
            
            messages.success(request, f"Datos cargados correctamente. Nuevos registros creados: {registros_creados}")
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")

        return redirect('cargar_datos_excel')  # Redirigir después de procesar

    return render(request, 'cargar_datos_excel.html')


#Seccion home Admin
def home(request):
    return render(request,'home.html')

def homeNeumatico(request):
    return render(request, 'areas/neumatico/homeneu.html')

def homeGPS(request):
    return render(request, 'areas/gps/homegps.html')

def homeDocumentacion(request):
    return render(request, 'areas/documento/homedoc.html')

def homeTaller(request):
    return render(request, 'areas/taller/hometaller.html')

def visual(request):
    return render(request, 'areas/visual/visual.html')
#Fin Admin

#Consulta (Arreglar)
@role_required(['Administrador', 'Visualizador'])
def consulta_vehiculo(request):
    query = request.GET.get('search', '')

    # Filtrar vehículos correctamente
    vehiculos = Vehiculo.objects.filter(
        Q(patente__icontains=query) | 
        Q(marca__nombre__icontains=query) |  # Acceder al campo "nombre" de la ForeignKey Marca
        Q(modelo__icontains=query)   
    )

    # Paginación: 15 vehículos por página
    paginator = Paginator(vehiculos, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'vehiculos': page_obj,
        'search': query
    }
    
    return render(request, 'Consulta/consulta.html', context)

# Vista para cargar los detalles del vehículo
def vehiculo_detalle(request, vehiculo_id):
    # Subconsulta para obtener el odómetro desde VehiculoAPI
    vehiculo_api_subquery = VehiculoAPI.objects.filter(
        placa=OuterRef('patente')  # Relacionar placa con patente
    ).values('odometro')[:1]  # Traemos solo el primer valor de odómetro

    # Obtener el vehículo y agregar el campo odómetro
    vehiculo = Vehiculo.objects.annotate(
        odometro=Subquery(vehiculo_api_subquery)
    ).get(id=vehiculo_id)

    # Pasar el vehículo a la plantilla
    return render(request, 'Consulta/seccion_detalles.html', {'vehiculo': vehiculo})

# Vista para cargar los documentos del vehículo
def vehiculo_documentos(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)
    return render(request, 'Consulta/seccion__documentos.html', {'vehiculo': vehiculo, 'documentos': documentos})

# Vista para cargar los hallazgos del vehículo
def vehiculo_hallazgos(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    hallazgos = HallazgoEmpresa.objects.filter(vehiculo=vehiculo)
    return render(request, 'Consulta/seccion__hallazgos.html', {'vehiculo': vehiculo, 'hallazgos': hallazgos})


# Vista para cargar los mantenimientos del vehículo
def vehiculo_mantenimientos(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    mantenimientos = Mantenimiento.objects.filter(vehiculo=vehiculo)
    return render(request, 'Consulta/seccion_mantenimientos.html', {'vehiculo': vehiculo, 'mantenimientos': mantenimientos})

# Vista para cargar las reparaciones del vehículo
def vehiculo_reparaciones(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    reparaciones = Reparacion.objects.filter(vehiculo=vehiculo)
    return render(request, 'Consulta/seccion_reparaciones.html', {'vehiculo': vehiculo, 'reparaciones': reparaciones})
#Fin Consulta

#Asiganciones

def asignar_empresa(request):
    # Obtener el término de búsqueda de la patente
    patente_buscar = request.GET.get('patente', '')  # Si no se pasa, por defecto estará vacío

    # Filtrar vehículos por patente si existe un término de búsqueda
    if patente_buscar:
        vehiculos_list = Vehiculo.objects.filter(patente__icontains=patente_buscar)  # Filtrar por patente (ignora mayúsculas/minúsculas)
    else:
        vehiculos_list = Vehiculo.objects.all()  # Obtener todos los vehículos si no se busca por patente
    
    empresas = Group.objects.all()  # Obtener todas las empresas

    # Paginación de 15 vehículos por página
    paginator = Paginator(vehiculos_list, 15)
    page_number = request.GET.get('page')  # Obtener el número de página
    page_obj = paginator.get_page(page_number)  # Obtener la página actual

    if request.method == "POST":
        form = AsignacionEmpresaForm(request.POST)
        if form.is_valid():
            vehiculo = form.cleaned_data['vehiculo']
            empresa = form.cleaned_data['empresa']
            vehiculo.empresa = empresa
            vehiculo.save()
            messages.success(request, f'Empresa asignada correctamente a {vehiculo.patente}')
            return redirect('asignar_empresa')
    else:
        form = AsignacionEmpresaForm()

    return render(request, 'asignacion/asignacion_empresa.html', {
        'vehiculos': page_obj,  # Pasar los vehículos paginados
        'form': form,
        'empresas': empresas,  # Asegúrate de pasar las empresas al contexto
        'patente_buscar': patente_buscar,  # Pasar el valor del campo de búsqueda para que se mantenga
    })

#Fin asiganciones

#Documentacion
@role_required(['Administrador', 'Visualizador'])
def Documentos(request): 
    # Obtener los valores de búsqueda y filtro
    query = request.GET.get('search', '')  # Búsqueda por patente
    tipo_filtro = request.GET.get('tipo', '')  # Filtro por tipo
    page_number = request.GET.get('page', 1)  # Número de página

    # Obtener todos los vehículos
    vehiculos = Vehiculo.objects.all()

    # Aplicar filtro por patente
    if query:
        vehiculos = vehiculos.filter(patente__icontains=query)

    # Aplicar filtro por tipo (buscando el ID del tipo)
    if tipo_filtro:
        try:
            tipo_obj = Tipo.objects.get(nombre=tipo_filtro)
            vehiculos = vehiculos.filter(tipo=tipo_obj.id)
        except Tipo.DoesNotExist:
            vehiculos = Vehiculo.objects.none()  # Si no existe el tipo, retorna lista vacía

    # Agregar la cantidad de documentos a cada vehículo
    for vehiculo in vehiculos:
        vehiculo.documentos_count = vehiculo.documento_set.count()

    # Configurar la paginación (10 elementos por página)
    paginator = Paginator(vehiculos, 10)
    page_obj = paginator.get_page(page_number)

    # Obtener opciones de tipos desde la base de datos
    tipos = Tipo.objects.values_list('nombre', 'nombre')  # Lista de tuplas (nombre, nombre)

    # Renderizar la plantilla con los datos
    return render(request, 'areas/documento/listadoc.html', {
        'page_obj': page_obj,  # Página actual
        'search': query,
        'tipo_filtro': tipo_filtro,
        'tipos': tipos,  # Opciones dinámicas desde la BD
    })

def subir_a_s3admin(archivo, carpeta):
    """
    Sube un archivo a S3 dentro de la carpeta especificada y devuelve la ruta S3.
    """
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        ruta_s3 = f"documento/{carpeta}/{archivo.name}"
        
        # Subir archivo
        s3_client.upload_fileobj(archivo, bucket_name, ruta_s3)
        print(f"Archivo subido correctamente: {ruta_s3}")
        return ruta_s3
    except Exception as e:
        print(f"Error al subir {archivo.name} a S3: {e}")
        return None

@role_required(['Administrador'])
def cargar_documentos(request, id):
    """
    Vista para cargar documentos de un vehículo y subirlos a S3 en subcarpetas específicas.
    """
    vehiculo = get_object_or_404(Vehiculo, id=id)

    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.vehiculo = vehiculo

            # Mapeo de documentos a subcarpetas dentro de 'documento/'
            subcarpetas = {
                'Mantencion_Preventiva': 'mantenciones',
                'Revision_Tecnica': 'revision_tecnica',
                'Permiso_Circulacion': 'permiso_circulacion',
                'SOAP': 'soap',
                'Padron': 'padron'
            }

            # Subir archivos a S3
            for campo, carpeta in subcarpetas.items():
                archivo = request.FILES.get(campo)
                if archivo:
                    s3_key = subir_a_s3admin(archivo, carpeta)
                    if s3_key:
                        setattr(documento, campo, s3_key)  # Guardar la ruta en el modelo
                    else:
                        messages.error(request, f"Error al subir {campo} a S3.")
                        return redirect('cargar_documentos', id=vehiculo.id)

            # Guardar el documento con las rutas S3
            documento.save()
            messages.success(request, "Documentos subidos correctamente.")
            return redirect('Documentos')

    else:
        form = DocumentoForm()

    return render(request, 'areas/documento/cargar_documentos.html', {
        'vehiculo': vehiculo,
        'form': form
    })

@role_required(['Administrador', 'Visualizador'])
def editar_documentos(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    if not documentos:
        return redirect('cargar_documentos', id=id)

    if request.method == 'POST':
        subcarpetas = {
            'Mantencion_Preventiva': 'mantenciones',
            'Revision_Tecnica': 'revision_tecnica',
            'Permiso_Circulacion': 'permiso_circulacion',
            'SOAP': 'soap',
            'Padron': 'padron'
        }

        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        for documento in documentos:
            form = DocumentoForm(request.POST, request.FILES, instance=documento)
            if form.is_valid():
                for campo, carpeta in subcarpetas.items():
                    if campo in request.FILES:  # Solo si el usuario subió un nuevo archivo en este campo
                        archivo_nuevo = request.FILES[campo]

                        s3_key = subir_a_s3admin(archivo_nuevo, carpeta)
                        if s3_key:
                            setattr(documento, campo, s3_key)
                        else:
                            messages.error(request, f"Error al subir {campo} a S3.")
                            return redirect('editar_documentos', id=vehiculo.id)

                documento.save()
            else:
                messages.error(request, "Error en el formulario de edición.")
                return redirect('editar_documentos', id=vehiculo.id)

        messages.success(request, "Documentos actualizados correctamente.")
        return redirect('Documentos')

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
    }
    return render(request, 'areas/documento/editar_documentos.html', context)

@role_required(['Administrador'])
def eliminar_documentosadm(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    if not documentos:
        messages.error(request, "No hay documentos asociados a este vehículo.")
        return redirect('cargar_documentos', vehiculo_id=vehiculo.id)

    if request.method == 'POST':
        # Conexión a S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Iterar sobre los documentos
        for documento in documentos:
            for campo in ['Mantencion_Preventiva', 'Revision_Tecnica', 'Permiso_Circulacion', 'SOAP', 'Padron']:
                checkbox_name = f"{campo}_{documento.id}"
                if request.POST.get(checkbox_name) == "delete":  # Si el checkbox fue marcado
                    archivo = getattr(documento, campo)
                    if archivo:  # Verifica que el archivo exista
                        try:
                            # Eliminar archivo de S3
                            s3_key = str(archivo)  # La URL del archivo en S3
                            s3_client.delete_object(
                                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                Key=s3_key
                            )
                            print(f"Archivo eliminado de S3: {s3_key}")
                            # Actualiza el campo en la base de datos a None
                            setattr(documento, campo, None)
                            documento.save()

                            # Si deseas eliminar completamente el documento de la base de datos
                            # documento.delete()

                        except Exception as e:
                            messages.error(request, f"Error al eliminar el archivo {campo} de S3: {e}")

        messages.success(request, "Documentos seleccionados eliminados correctamente.")
        return redirect('editar_documentos', vehiculo_id=vehiculo.id)

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
    }
    return render(request, 'areas/documento/eliminar_documentosadm.html', context)

def cargar_fechas_documentos(request):
    usuario = request.user  # Usuario autenticado
    grupo_usuario = usuario.group  # Obtener el grupo del usuario

    if not grupo_usuario:
        return JsonResponse({"error": "El usuario no pertenece a ningún grupo."}, status=403)

    if request.method == 'POST' and 'archivo' in request.FILES:
        archivo_excel = request.FILES['archivo']
        df = pd.read_excel(archivo_excel, engine='openpyxl')  # Leer el archivo Excel

        documentos_actualizados = 0
        errores = []

        for index, row in df.iterrows():
            try:
                # Buscar el vehículo por patente, filtrando solo los que pertenecen al grupo del usuario
                vehiculo = Vehiculo.objects.filter(patente=row['Patente'], empresa=grupo_usuario).first()
                if not vehiculo:
                    errores.append(f"Vehículo con patente {row['Patente']} no encontrado o no pertenece a su empresa.")
                    continue

                # Buscar el documento asociado al vehículo o crearlo si no existe
                doc, _ = Documento.objects.get_or_create(vehiculo=vehiculo)

                # Actualizar fechas si están presentes en el Excel
                doc.fecha_vencimiento_mantencion = row.get('Fecha Mantención', doc.fecha_vencimiento_mantencion)
                doc.fecha_vencimiento_revision = row.get('Fecha Revisión', doc.fecha_vencimiento_revision)
                doc.fecha_vencimiento_permiso = row.get('Fecha Permiso Circulación', doc.fecha_vencimiento_permiso)
                doc.fecha_vencimiento_soap = row.get('Fecha SOAP', doc.fecha_vencimiento_soap)
                doc.fecha_vencimiento_padron = row.get('Fecha Padrón', doc.fecha_vencimiento_padron)

                doc.save()
                documentos_actualizados += 1

            except Exception as e:
                errores.append(f"Error en fila {index + 1}: {str(e)}")

        return JsonResponse({
            "mensaje": f"Se actualizaron {documentos_actualizados} documentos.",
            "errores": errores
        })

    return render(request, 'empresa/cargar_fechas.html')

@role_required(['Administrador'])
def lista_eliminarADM(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
    }

    return render(request, 'areas/documento/lista_eliminar.html',context)

def subir_a_s3MTTO(archivo, carpeta="mantenimientos"):
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        # Generar un nombre único para el archivo usando UUID
        extension = archivo.name.split('.')[-1]  # Obtiene la extensión
        nombre_archivo = f"{carpeta}/{uuid.uuid4()}.{extension}"
        
        # Subir el archivo a S3
        s3_client.upload_fileobj(archivo, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)
        print(f"Archivo subido correctamente: {nombre_archivo}")
        return nombre_archivo  # Retorna la key (ruta) del archivo en S3
    except Exception as e:
        print(f"Error al subir archivo a S3: {e}")
        return None

@role_required(['Administrador'])
def crear_mantenimiento(request):
    if request.method == 'POST':
        form = MantenimientoForm(request.POST, request.FILES)
        if form.is_valid():
            # Creamos el objeto sin guardarlo inmediatamente
            mantenimiento = form.save(commit=False)
            
            # Verificamos si se envió un archivo para el campo respaldo_mtto
            respaldo_file = request.FILES.get('respaldo_mtto')
            if respaldo_file:
                s3_key = subir_a_s3MTTO(respaldo_file, carpeta="mantenimientos")
                if s3_key:
                    # Asignamos la key retornada al campo respaldo_mtto
                    mantenimiento.respaldo_mtto = s3_key
                else:
                    messages.error(request, "Error al subir el respaldo a S3")
                    return redirect('crear_mantenimiento')  # O renderiza el formulario con error
            
            # Guardamos el objeto Mantenimiento
            mantenimiento.save()
            messages.success(request, "Mantenimiento creado exitosamente")
            return redirect('homeDocumentacion')
    else:
        search = request.GET.get('search', '')
        form = MantenimientoForm()
        if search:
            form.fields['vehiculo'].queryset = Vehiculo.objects.filter(patente__icontains=search)
    
    return render(request, 'areas/documento/crear_mantenimiento.html', {'form': form})

@role_required(['Administrador', 'Visualizador'])
def listado_vehiculos(request):
    # Obtener el término de búsqueda
    search = request.GET.get('search', '')
    
    # Filtrar vehículos por patente si existe la búsqueda
    if search:
        vehiculos = Vehiculo.objects.filter(patente__icontains=search)
    else:
        vehiculos = Vehiculo.objects.all()

    return render(request, 'areas/documento/lista_vehiculos.html', {
        'vehiculos': vehiculos,
        'search': search,
    })

@role_required(['Administrador', 'Visualizador'])
def historial_mantenimiento(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    mantenimientos = Mantenimiento.objects.filter(vehiculo=vehiculo).order_by('-fecha_mtto')
    
    # Crear el cliente de S3 una sola vez para mayor eficiencia
    s3_client = boto3.client(
        's3',
        region_name=settings.AWS_S3_REGION_NAME,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    tiempo_expiracion = 240  # Tiempo de expiración en segundos

    # Iteramos sobre cada mantenimiento para generar la URL temporal si existe respaldo_mtto
    for mantenimiento in mantenimientos:
        if mantenimiento.respaldo_mtto:
            archivo_s3 = str(mantenimiento.respaldo_mtto.name)
            try:
                mantenimiento.url_temporal = s3_client.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
                    ExpiresIn=tiempo_expiracion
                )
            except Exception as e:
                print(f"Error al generar URL temporal para mantenimiento {mantenimiento.id}: {e}")
                mantenimiento.url_temporal = None
        else:
            mantenimiento.url_temporal = None

    return render(request, 'areas/documento/historial_vehiculo.html', {
        'vehiculo': vehiculo,
        'mantenimientos': mantenimientos,
    })

def mantenimiento_view(request):
    # Obtener todos los vehículos y sus mantenimientos
    vehiculos = VehiculoAPI.objects.all()
    mantenimiento_data = []

    for vehiculo in vehiculos:
        # Buscar el último mantenimiento del vehículo
        mantenimiento = Mantenimiento.objects.filter(vehiculo=vehiculo.vehiculo).order_by('-fecha_mtto').first()

        if mantenimiento:
            # Calcular la diferencia entre el odómetro y el próximo mantenimiento
            odometro_actual = vehiculo.odometro
            proximo_mantenimiento_km = mantenimiento.proximo_mantenimiento_km
            diferencia_km = proximo_mantenimiento_km - odometro_actual

            # Determinar el estado según la diferencia de kilómetros
            if diferencia_km > 3000:
                estado = 'Vigente'
            elif diferencia_km <= 3000 and diferencia_km > 0:
                estado = 'Por Vencer'
            else:
                estado = 'Vencida'

            mantenimiento_data.append({
                'patente': vehiculo.placa,
                'kilometraje_mtto': mantenimiento.kilometraje_mtto,
                'fecha_mtto': mantenimiento.fecha_mtto,
                'servicio_realizado': mantenimiento.servicio_realizado,
                'odometro_actual': odometro_actual,
                'proximo_mantenimiento_km': proximo_mantenimiento_km,
                'proximo_servicio': mantenimiento.proximo_servicio,
                'estado': estado,
            })

    return render(request, 'areas/documento/mantenimiento.html', {'mantenimiento_data': mantenimiento_data})
#Fin Documentacion

#Conductores
@role_required(['Administrador', 'Visualizador'])
def listar_conductores(request):
    conductores = Conductor.objects.all()
    
    # Implementación de la paginación
    paginator = Paginator(conductores, 15)  # 15 conductores por página
    page_number = request.GET.get('page')  # Obtiene el número de página desde la URL
    page_obj = paginator.get_page(page_number)
    
    # Lógica para determinar el estado de la licencia de cada conductor
    for conductor in page_obj:
        # Verificar si el conductor tiene licencias asociadas
        licencia = conductor.licencias.first()  # Se obtiene la primera licencia asociada al conductor
        if licencia and licencia.archivo:
            # Verificamos la fecha de vencimiento de la licencia
            fecha_vencimiento = conductor.FechaVencimientoLicencia

            # Verifica si la licencia está vencida, por vencer o vigente
            if fecha_vencimiento < timezone.now().date():
                conductor.estado_licencia = "Vencida"
            elif fecha_vencimiento <= (timezone.now().date() + timedelta(days=15)):
                conductor.estado_licencia = "Por Vencer"
            else:
                conductor.estado_licencia = "Vigente"
        else:
            conductor.estado_licencia = "Sin Respaldo Licencia"
    
    return render(request, 'areas/documento/conductores/list.html', {'page_obj': page_obj})

@role_required(['Administrador'])
def crear_conductor(request):
    if request.method == 'POST':
        form = ConductorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Conductor agregado exitosamente.")
            return redirect('list')
    else:
        form = ConductorForm()
    return render(request, 'areas/documento/conductores/form.html', {'form': form})


@role_required(['Administrador'])
def subir_licencia(request, conductor_id):
    conductor = get_object_or_404(Conductor, id=conductor_id)
    
    if request.method == 'POST':
        archivos = request.FILES.getlist('archivo_licencia')

        if archivos:
            for archivo in archivos:
                url_archivo = subir_a_s3(archivo, "licencias")

                if url_archivo:
                    # Guardar en la base de datos
                    LicenciaConductor.objects.create(
                        conductor=conductor,
                        archivo=url_archivo
                    )
                else:
                    messages.error(request, "Hubo un error al subir un archivo.")

            messages.success(request, "Licencia subida exitosamente.")
            return redirect('list')  # Redirige a la lista de conductores
        else:
            messages.error(request, "Por favor, sube al menos un archivo de licencia.")
    
    return render(request, 'areas/documento/conductores/subir_licencia.html', {'conductor': conductor})

@role_required(['Administrador'])
def editar_licencia(request, conductor_id):
    conductor = get_object_or_404(Conductor, id=conductor_id)
    
    # Obtener las licencias existentes del conductor
    licencias = LicenciaConductor.objects.filter(conductor=conductor)

    if request.method == 'POST':
        archivos = request.FILES.getlist('archivo_licencia')
        
        if archivos:
            # Si no se marca la opción de mantener las licencias, eliminamos las existentes
            if not request.POST.get('mantener_licencias'):
                # Convertir el QuerySet a una lista de diccionarios para imprimir los datos (para depuración)
                licencias_eliminadas = list(licencias.values())
                print("Licencias que se eliminarán:", licencias_eliminadas)
                
                # Instanciar el cliente de S3
                s3_client = boto3.client(
                    's3',
                    region_name=settings.AWS_S3_REGION_NAME,
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
                )
                
                # Recorrer cada licencia y eliminar su archivo de S3
                for licencia in licencias:
                    s3_key = licencia.archivo.name  # Obtiene la clave (key) del archivo en S3
                    try:
                        s3_client.delete_object(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                            Key=s3_key
                        )
                        print(f"Archivo eliminado de S3: {s3_key}")
                    except Exception as e:
                        print(f"Error al eliminar {s3_key} de S3: {e}")
                
                # Luego, eliminar los registros de la base de datos
                licencias.delete()
            
            # Para cada archivo nuevo, lo subimos a S3 y guardamos su ruta en la base de datos
            for archivo in archivos:
                url_archivo = subir_a_s3(archivo, "licencias")
                if url_archivo:
                    LicenciaConductor.objects.create(
                        conductor=conductor,
                        archivo=url_archivo
                    )
                else:
                    messages.error(request, "Hubo un error al subir uno de los archivos a S3.")
                    return redirect('editar_licencia', conductor_id=conductor.id)
            
            messages.success(request, "Licencias actualizadas exitosamente.")
            return redirect('list')  # Redirige a la lista de conductores
        else:
            messages.error(request, "Por favor, sube al menos un archivo de licencia.")
    
    return render(request, 'areas/documento/conductores/editar_licencia.html', {
        'conductor': conductor,
        'licencias': licencias,  # Se pasan las licencias existentes al template
    })


@role_required(['Administrador'])
def editar_conductor(request, pk):
    conductor = get_object_or_404(Conductor, pk=pk)
    if request.method == 'POST':
        form = ConductorForm(request.POST, instance=conductor)
        if form.is_valid():
            form.save()
            messages.success(request, "Conductor actualizado correctamente.")
            return redirect('list')
    else:
        form = ConductorForm(instance=conductor)
    return render(request, 'areas/documento/conductores/form.html', {'form': form})

@role_required(['Administrador'])
def eliminar_conductor(request, pk):
    conductor = get_object_or_404(Conductor, pk=pk)
    if request.method == 'POST':
        conductor.delete()
        messages.success(request, "Conductor eliminado correctamente.")
        return redirect('list')
    return render(request, 'areas/documento/conductores/confirm_delete.html', {'conductor': conductor})

@role_required(['Administrador'])
def importar_conductores(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "No se ha subido ningún archivo.")
            return redirect('import')  

        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):  # Limitar a las 5 columnas necesarias
                if row[0]:  # Evita procesar filas vacías
                    nombre, rut, telefono, fecha_vencimiento, empresa_nombre = row

                    # Buscar la empresa por su nombre en la base de datos
                    try:
                        empresa = Group.objects.get(name=empresa_nombre)
                    except Group.DoesNotExist:
                        messages.error(request, f"Empresa '{empresa_nombre}' no encontrada.")
                        continue

                    # Asegurarse de que la fecha es un objeto de tipo datetime y convertirlo si es necesario
                    try:
                        if isinstance(fecha_vencimiento, datetime):
                            fecha_vencimiento = fecha_vencimiento.date()  # Si es un objeto datetime, tomar solo la fecha
                        else:
                            # Si la fecha no es un objeto datetime, intentar convertirla
                            fecha_vencimiento = datetime.strptime(str(fecha_vencimiento), '%d-%m-%Y').date()

                    except Exception as e:
                        messages.error(request, f"Fecha de vencimiento inválida para {nombre}: {str(e)}")
                        continue

                    Conductor.objects.create(
                        nombre=nombre,
                        rut=rut,
                        telefono=telefono,
                        FechaVencimientoLicencia=fecha_vencimiento,
                        empresa=empresa
                    )
            messages.success(request, "Conductores importados exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al importar: {str(e)}")

    return render(request, 'areas/documento/conductores/importar.html')  # Renderiza el formulario


def licencia_detalle(request, licencia_id):
    licencia = get_object_or_404(LicenciaConductor, id=licencia_id)

    # Obtener el nombre del archivo como cadena
    archivo_s3 = str(licencia.archivo.name)  # Asegurar que es una cadena
    print(f"Archivo S3: {archivo_s3}")

    tiempo_expiracion = 240

    s3_client = boto3.client(
        's3',
        region_name=settings.AWS_S3_REGION_NAME,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )

    # Generar URL temporal
    try:
        url_temporal = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
            ExpiresIn=tiempo_expiracion
        )
        #print(f"URL temporal generada: {url_temporal}")   
    except Exception as e:
        #print(f"Error al generar el enlace temporal: {e}")
        url_temporal = None

    return render(request, 'areas/documento/conductores/licencia_detalle.html', {'licencia': licencia, 'url_temporal': url_temporal})

#Fin Conductores

#Inicio Neumatico

def subir_a_s3_hall(archivo, carpeta="evidencias"):
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Generar un nombre único para el archivo
        extension = archivo.name.split('.')[-1]  # Obtener la extensión
        nombre_archivo = f"{carpeta}/{uuid.uuid4()}.{extension}"  # Carpeta + UUID + extensión

        # Subir archivo a S3
        s3_client.upload_fileobj(archivo, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)

        print(f"Archivo subido correctamente: {nombre_archivo}")
        return nombre_archivo  # Retorna la ruta (key) del archivo en S3
    except Exception as e:
        print(f"Error al subir archivo a S3: {e}")
        return None

@role_required(['Administrador'])
def crear_hallazgo(request):
    if request.method == 'POST':
        form = HallazgoForm(request.POST, request.FILES)
        if form.is_valid():
            hallazgo = form.save(commit=False)
            
            # Verificar si se envió un archivo en el campo "evidencia"
            evidencia_file = request.FILES.get('evidencia')
            if evidencia_file:
                s3_key = subir_a_s3_hall(evidencia_file, carpeta="evidencias")
                if s3_key is not None:
                    # Asignamos la ruta devuelta al campo 'evidencia'.
                    # De esta forma, en la base de datos se almacenará, por ejemplo:
                    # "evidencias/7b3f3e2c-1b34-4f70-bf76-xxxxx.jpg"
                    hallazgo.evidencia = s3_key
                else:
                    messages.error(request, "Error al subir la evidencia a S3.")
                    return redirect('crear_hallazgo')
            
            hallazgo.save()  # Guardamos el objeto para tener su ID y demás
            
            # Enviar notificación a los miembros del grupo (según tu lógica)
            asunto = f"Nuevo Hallazgo - {hallazgo.hallazgo}"
            mensaje = f"Se ha creado un nuevo hallazgo con ID {hallazgo.id}."
            enviar_notificacion_grupo(hallazgo, asunto, mensaje, tipo='creacion')

            messages.success(request, 'Hallazgo creado exitosamente')
            return redirect('listar_hallazgoemp')
    else:
        form = HallazgoForm()

    return render(request, 'areas/neumatico/crear_hallazgo.html', {'form': form})

@role_required(['Administrador', 'Visualizador'])
def listar_hallazgoemp(request):
    # Obtener todos los hallazgos sin ningún filtro
    hallazgos = HallazgoEmpresa.objects.all()

    return render(request, 'areas/neumatico/lista_hallazgo.html', {'hallazgos': hallazgos})

@role_required(['Administrador', 'Visualizador'])
def detalle_hallazgoemp(request, hallazgo_id):
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)
    cierre = Cierre.objects.filter(hallazgo=hallazgo).first()  # Obtiene el cierre si existe

    url_temporal = None  # Inicializamos la variable

    # Verificamos si hay evidencia cargada en el hallazgo
    if hallazgo.evidencia:
        archivo_s3 = str(hallazgo.evidencia.name)  # Obtener la key del archivo en S3
        tiempo_expiracion = 240  # Tiempo de expiración en segundos

        # Creamos el cliente de S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Generamos la URL temporal
        try:
            url_temporal = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
                ExpiresIn=tiempo_expiracion
            )
        except Exception as e:
            print(f"Error al generar URL temporal para evidencia: {e}")
            url_temporal = None

    return render(request, 'areas/neumatico/detalle_hallazgo.html', {
        'hallazgo': hallazgo,
        'cierre': cierre,  # Pasamos el cierre a la plantilla (si existe)
        'url_temporal': url_temporal  # Pasamos la URL temporal para la evidencia
    })

@role_required(['Administrador'])
def cerrar_hallazgoemp(request, hallazgo_id):
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)

    # Verificar si el estado es 'Pendiente' para poder cerrarlo
    if hallazgo.estado == 'Pendiente':
        hallazgo.estado = 'Cerrado'  # Cambiar el estado a Cerrado
        hallazgo.descripcion_cierre = "Descripción del cierre aquí"  # Puedes poner la lógica para capturar esta descripción si es necesario
        hallazgo.save()

    return redirect('detalle_hallazgoemp', hallazgo_id=hallazgo.id)  # Redirige a la página de detalles del hallazgo

@role_required(['Administrador'])
def reabrir_hallazgo(request, hallazgo_id):
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)

    # Solo reabrir si está cerrado
    if hallazgo.estado == 'Cerrado':
        hallazgo.estado = 'Pendiente'  # Cambiar el estado a Pendiente
        hallazgo.descripcion_cierre = None  # Limpiar la descripción de cierre
        hallazgo.evidencia_cierre = None  # Limpiar la evidencia de cierre
        hallazgo.documento_cierre = None  # Limpiar el documento de cierre
        hallazgo.save()

    return redirect('detalle_hallazgoemp', hallazgo_id=hallazgo.id)

@role_required(['Administrador'])
def add_comunicacion(request, hallazgo_id):
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)

    if request.method == 'POST':
        form = ComunicacionForm(request.POST, request.FILES)
        if form.is_valid():
            comunicacion = form.save(commit=False)
            comunicacion.hallazgo = hallazgo
            comunicacion.usuario = request.user
            comunicacion.save()
            return redirect('detalle_hallazgoemp', hallazgo_id=hallazgo.id)
    else:
        form = ComunicacionForm()

    return render(request, 'areas/neumatico/comunicacion_form.html', {'form': form, 'hallazgo': hallazgo})

#Fin Neumatico

#Empresa Hallazgo

@role_required(['Empresa'])
def empresa_hallazgos(request):
    # Obtener el grupo del usuario autenticado
    grupo_usuario = request.user.group  # Suponiendo que el usuario pertenece a un único grupo
    if not grupo_usuario:
        return render(request, 'empresa/hallazgos_list.html', {'hallazgos': [], 'grupo': None})
    
    # Filtrar los hallazgos del grupo del usuario
    hallazgos = Hallazgo.objects.filter(grupo__name=grupo_usuario.name).order_by('-fecha_inspeccion')
    
    # Paginación (10 hallazgos por página)
    paginator = Paginator(hallazgos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'empresa/hallazgos_list.html', {
        'hallazgos': page_obj,
        'grupo': grupo_usuario
    })

#Fin Empresa Hallazgo

#Empresa Vehiculos
@role_required(['Empresa'])
def vehiculos_del_grupo(request):
    # Obtener el grupo del usuario actual
    grupo_usuario = request.user.group

    # Obtener el término de búsqueda y tipo de filtro (si existen)
    search = request.GET.get('search', '')

    # Filtrar los vehículos que pertenecen al mismo grupo y aplicar el filtro de búsqueda
    vehiculos = Vehiculo.objects.filter(empresa=grupo_usuario)

    # Aplicar filtro de búsqueda por patente si se proporciona
    if search:
        vehiculos = vehiculos.filter(patente__icontains=search)

    # Agregar la cantidad de documentos a cada vehículo
    for vehiculo in vehiculos:
        vehiculo.documentos_count = vehiculo.documento_set.count()  # Cuenta los documentos asociados al vehículo

    # Ordenar los vehículos por patente
    vehiculos = vehiculos.order_by('patente')

    # Implementación de paginación
    paginator = Paginator(vehiculos, 10)  # 10 vehículos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'empresa/lista_vehiculosemp.html', {
        'page_obj': page_obj,
        'search': search,  # Pasar el término de búsqueda para mantenerlo en el formulario
    })



@role_required(['Empresa'])
def cargar_documentosemp(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.vehiculo = vehiculo

            # Mapeo de documentos a subcarpetas dentro de 'documento/'
            subcarpetas = {
                'Mantencion_Preventiva': 'mantenciones',
                'Revision_Tecnica': 'revision_tecnica',
                'Permiso_Circulacion': 'permiso_circulacion',
                'SOAP': 'soap',
                'Padron': 'padron'
            }

            # Subir archivos a S3
            for campo, carpeta in subcarpetas.items():
                archivo = request.FILES.get(campo)
                if archivo:
                    s3_key = subir_a_s3admin(archivo, carpeta)
                    if s3_key:
                        setattr(documento, campo, s3_key)  # Guardar la ruta en el modelo
                    else:
                        messages.error(request, f"Error al subir {campo} a S3.")
                        return redirect('vehiculos_del_grupo')  # Redirige a la lista de vehículos
            documento.save()
            messages.success(request, "Documentos subidos correctamente.")
            return redirect('vehiculos_del_grupo')
    else:
        form = DocumentoForm()
    
    return render(request, 'empresa/cargar_documentosemp.html', {'form': form, 'vehiculo': vehiculo})

@role_required(['Empresa'])
def editar_documentosemp(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    if not documentos:
        return redirect('cargar_documentosemp', vehiculo_id=vehiculo.id)

    # Inicializamos el formulario con los datos del primer documento asociado
    form = DocumentoForm(instance=documentos.first())

    # Definimos las subcarpetas
    subcarpetas = {
        'Mantencion_Preventiva': 'mantenciones',
        'Revision_Tecnica': 'revision_tecnica',
        'Permiso_Circulacion': 'permiso_circulacion',
        'SOAP': 'soap',
        'Padron': 'padron'
    }

    if request.method == 'POST':
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        for documento in documentos:
            form = DocumentoForm(request.POST, request.FILES, instance=documento)
            if form.is_valid():
                for campo, carpeta in subcarpetas.items():
                    if campo in request.FILES:  # Si se subió un nuevo archivo
                        nuevo_archivo = request.FILES[campo]

                        s3_key = subir_a_s3admin(nuevo_archivo, carpeta)
                        if s3_key:
                            setattr(documento, campo, s3_key)  # Asignar la nueva URL del archivo en S3 al modelo
                        else:
                            messages.error(request, f"Error al subir {campo} a S3.")
                            return redirect('editar_documentosemp', vehiculo_id=vehiculo.id)

                # 🔹 Guardar el documento después de agregar el nuevo archivo
                documento.save()

        messages.success(request, "Documentos actualizados correctamente.")
        return redirect('vehiculos_del_grupo')

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
        'form': form,
    }

    return render(request, 'empresa/editar_documentosemp.html', context)

@role_required(['Empresa'])
def lista_eliminar(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
    }

    return render(request, 'empresa/lista_eliminar.html',context)

def eliminar_documentosemp(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    documentos = Documento.objects.filter(vehiculo=vehiculo)

    if not documentos:
        messages.error(request, "No hay documentos asociados a este vehículo.")
        return redirect('cargar_documentosemp', vehiculo_id=vehiculo.id)

    if request.method == 'POST':
        # Conexión a S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Iterar sobre los documentos
        for documento in documentos:
            for campo in ['Mantencion_Preventiva', 'Revision_Tecnica', 'Permiso_Circulacion', 'SOAP', 'Padron']:
                checkbox_name = f"{campo}_{documento.id}"
                if request.POST.get(checkbox_name) == "delete":  # Si el checkbox fue marcado
                    archivo = getattr(documento, campo)
                    if archivo:  # Verifica que el archivo exista
                        try:
                            # Eliminar archivo de S3
                            s3_key = str(archivo)  # La URL del archivo en S3
                            s3_client.delete_object(
                                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                Key=s3_key
                            )
                            print(f"Archivo eliminado de S3: {s3_key}")
                            # Actualiza el campo en la base de datos a None
                            setattr(documento, campo, None)
                            documento.save()

                            # Si deseas eliminar completamente el documento de la base de datos
                            # documento.delete()

                        except Exception as e:
                            messages.error(request, f"Error al eliminar el archivo {campo} de S3: {e}")

        messages.success(request, "Documentos seleccionados eliminados correctamente.")
        return redirect('editar_documentosemp', vehiculo_id=vehiculo.id)

    context = {
        'vehiculo': vehiculo,
        'documentos': documentos,
    }
    return render(request, 'empresa/eliminar_documentosemp.html', context)



@role_required(['Empresa'])
def lista_documentos(request):
    grupo_usuario = request.user.group
    vehiculos = Vehiculo.objects.filter(empresa=grupo_usuario)
    documentos = Documento.objects.filter(vehiculo__in=vehiculos)
    today = timezone.now().date()
    vehiculos_estado = []

    for vehiculo in vehiculos:
        estado = {
            'patente': vehiculo.patente,
            'revision': 'Sin Documento',
            'padron': 'Sin Documento',
            'mantenimiento': 'Sin Documento',
            'soap': 'Sin Documento',
            'permiso_circulacion': 'Sin Documento',
        }

        vehiculo_documentos = documentos.filter(vehiculo=vehiculo)

        if vehiculo_documentos.exists():
            for doc in vehiculo_documentos:
                # Función helper para determinar el estado
                def determinar_estado(fecha):
                    if not fecha:
                        return 'Sin Documento'
                    
                    dias_para_vencer = (fecha - today).days
                    
                    if dias_para_vencer <= 0:
                        return 'Vencido'
                    elif dias_para_vencer <= 30:
                        return 'Por Vencer'
                    else:
                        return 'Vigente'

                # Aplicar la lógica a cada tipo de documento
                estado['revision'] = determinar_estado(doc.fecha_vencimiento_revision)
                estado['padron'] = determinar_estado(doc.fecha_vencimiento_padron)
                estado['mantenimiento'] = determinar_estado(doc.fecha_vencimiento_mantencion)
                estado['soap'] = determinar_estado(doc.fecha_vencimiento_soap)
                estado['permiso_circulacion'] = determinar_estado(doc.fecha_vencimiento_permiso)

        vehiculos_estado.append(estado)

    return render(request, 'empresa/lista_documentos.html', {'vehiculos_estado': vehiculos_estado})
#Fin Empresa Vehiculos

#Inicio Taller Admin
@role_required(['Administrador'])
def crear_asignacion(request):
    if request.method == 'POST':
        form = AsignacionVehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('homeTaller')  # Cambia por tu vista de redirección
    else:
        form = AsignacionVehiculoForm()
    return render(request, 'areas/taller/crear_asignacion.html', {'form': form})

@role_required(['Administrador', 'Visualizador'])
def listar_asignaciones(request):
    # Obtener todas las asignaciones sin aplicar filtros
    asignaciones = Asignacion_taller.objects.all()

    return render(request, 'areas/taller/listar_asignaciones.html', {'asignaciones': asignaciones})

@role_required(['Administrador', 'Visualizador'])
def listar_mantencion_reparacion(request):
    # Filtrar asignaciones de tipo 'Mantención Preventiva', 'Mantención Correctiva', y 'Reparación'
    asignaciones = Asignacion_taller.objects.filter(tipo__in=['mantencion_preventiva', 'mantencion_correctiva', 'reparacion'])

    return render(request, 'areas/taller/listar_asignaciones.html', {'asignaciones': asignaciones})

@role_required(['Administrador', 'Visualizador'])
def listar_asignaciones_empresa(request):
    # Filtrar asignaciones de tipo 'Asignación Empresa'
    asignaciones = Asignacion_taller.objects.filter(tipo='asignacion_empresa')

    return render(request, 'areas/taller/listar_asignaciones_empresa.html', {'asignaciones': asignaciones})

@role_required(['Administrador', 'Visualizador'])
def unidad_aceptada_list(request):
    # Obtener todas las unidades aceptadas sin ningún filtro
    unidades = UnidadAceptada.objects.all()
    
    return render(request, 'areas/taller/unidad_aceptada_list.html', {'unidades': unidades})

@role_required(['Administrador', 'Visualizador'])
def unidades_pendientes(request):
    unidades = UnidadAceptada.objects.filter(estado='Pendiente')
    return render(request, 'areas/taller/unidades_pendientes.html', {'unidades': unidades})

@role_required(['Administrador', 'Visualizador'])
def unidades_en_proceso(request):
    unidades = UnidadAceptada.objects.filter(estado='En Proceso')
    return render(request, 'areas/taller/unidades_en_proceso.html', {'unidades': unidades})

@role_required(['Administrador', 'Visualizador'])
def unidades_reparadas(request):
    unidades = UnidadAceptada.objects.filter(estado='Reparada')
    return render(request, 'areas/taller/unidades_reparadas.html', {'unidades': unidades})

@role_required(['Administrador', 'Visualizador'])
def detalle_reparacion(request, id):
    unidad = get_object_or_404(UnidadAceptada, id=id)

    url_temporal = None  # Inicializamos la variable

    # Verificamos si hay evidencia cargada en el hallazgo
    if unidad.registro:
        archivo_s3 = str(unidad.registro.name)  # Obtener la key del archivo en S3
        tiempo_expiracion = 240  # Tiempo de expiración en segundos

        # Creamos el cliente de S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Generamos la URL temporal
        try:
            url_temporal = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
                ExpiresIn=tiempo_expiracion
            )
            print(url_temporal)
        except Exception as e:
            print(f"Error al generar URL temporal para evidencia: {e}")
            url_temporal = None

    return render(request, 'areas/taller/detalle_reparacion.html', {'unidad': unidad, 'url_temporal': url_temporal})

@role_required(['Administrador', 'Visualizador'])
def marcar_como_reparada(request, unidad_id):
    # Asegúrate de que la unidad existe
    unidad = UnidadAceptada.objects.get(id=unidad_id)
    
    # Actualiza el estado de la unidad a "Reparada"
    unidad.estado = 'Reparada'
    unidad.save()
    
    # Redirige de vuelta a la página de unidades reparadas
    return redirect('unidades_reparadas')

def chat_view(request):
    groups = Group.objects.all()
    selected_group_id = request.GET.get('group_id')
    selected_group = Group.objects.get(id=selected_group_id) if selected_group_id else None
    messages = Message.objects.filter(group=selected_group).order_by('timestamp') if selected_group else []

    if request.method == 'POST':
        content = request.POST.get('content')
        group_id = request.POST.get('group_id')
        if content and group_id:
            group = Group.objects.get(id=group_id)
            Message.objects.create(
                sender=request.user,
                group=group,
                content=content,
                timestamp=timezone.now()
            )
            return redirect('chat')  # Redirigir para evitar reenvío de formulario

    return render(request, 'chat/chat.html', {'groups': groups, 'messages': messages, 'selected_group': selected_group})



#Fin Taller Admin

#Chat

def admin_chat_view(request):
    if not request.user.role or request.user.role.name != 'Administrador':
        return redirect('user_chat')  # Redirigir a la vista de usuario si no es admin

    groups = Group.objects.all()
    selected_group_id = request.GET.get('group_id')
    selected_group = Group.objects.get(id=selected_group_id) if selected_group_id else None
    messages = Message.objects.filter(group=selected_group).order_by('timestamp') if selected_group else []

    if request.method == 'POST':
        content = request.POST.get('content')
        group_id = request.POST.get('group_id')
        if content and group_id:
            group = Group.objects.get(id=group_id)
            message = Message.objects.create(
                sender=request.user,
                group=group,
                content=content,
                timestamp=timezone.now()
            )
            # Devolver los datos del nuevo mensaje en formato JSON
            return JsonResponse({
                'success': True,
                'message': {
                    'sender': message.sender.username,
                    'content': message.content,
                    'timestamp': message.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                }
            })

    return render(request, 'chat/admin_chat.html', {'groups': groups, 'messages': messages, 'selected_group': selected_group})

@role_required(['Empresa','Taller'])
def user_chat_view(request):
    user_group = request.user.group
    messages = Message.objects.filter(group=user_group).order_by('timestamp') if user_group else []
    
    # Marcar los mensajes como leídos
    if messages:
        Message.objects.filter(group=user_group, read=False).update(read=True)

    # Contar los mensajes no leídos
    unread_count = Message.objects.filter(group=user_group, read=False).count() if user_group else 0

    if request.method == 'POST':
        content = request.POST.get('content')
        if content and user_group:
            # Crear el mensaje
            message = Message.objects.create(
                sender=request.user,
                group=user_group,
                content=content,
                timestamp=timezone.now()
            )
            # Devolver los datos del mensaje creado como respuesta JSON
            return JsonResponse({
                'username': message.sender.username,
                'content': message.content,
                'timestamp': message.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            })

    return render(request, 'chat/user_chat.html', {'messages': messages, 'user_group': user_group, 'unread_count': unread_count})

#Fin Chat

#Inicio Taller usuario
@role_required(['Taller'])
def actualizar_estado(request, asignacion_id):
    asignacion = AsignacionVehiculo.objects.get(id=asignacion_id)
    if request.method == "POST":
        form = ActualizarEstadoForm(request.POST, instance=asignacion)
        if form.is_valid():
            form.save()
            return redirect('listar_asignaciones')
    else:
        form = ActualizarEstadoForm(instance=asignacion)
    return render(request, 'taller/actualizar_estado.html', {'form': form, 'asignacion': asignacion})

@role_required(['Taller'])
def gestionar_asignaciones(request):
    grupo_usuario = request.user.group

    if grupo_usuario:
        asignaciones = Asignacion_taller.objects.filter(taller=grupo_usuario)
    else:
        asignaciones = Asignacion_taller.objects.none()

    if request.method == 'POST':
        asignacion_id = request.POST.get('asignacion_id')
        estado = request.POST.get('estado')
        comentario_rechazo = request.POST.get('comentario_rechazo', '')
        fecha_retiro = request.POST.get('fecha_retiro', '')
        comentario = request.POST.get('comentario', '')

        asignacion = get_object_or_404(Asignacion_taller, id=asignacion_id)

        if estado == 'Aceptada':
            if not fecha_retiro or not comentario:
                messages.error(request, "Debe proporcionar todos los campos requeridos.")
                return redirect('gestionar_asignaciones')

            # Crear respuesta de asignación
            RespuestaAsignacion_taller.objects.create(
                asignacion=asignacion,
                usuario=request.user,
                estado=estado,
                comentario_rechazo='',
                fecha_retiro=fecha_retiro,
                comentario=comentario,
            )

            # Crear registro en UnidadAceptada
            UnidadAceptada.objects.create(
                patente=asignacion.patente,
                taller=asignacion.taller,
                fecha_respuesta=asignacion.respuestas.last().fecha_respuesta,
                fecha_retiro=fecha_retiro,
            )

            messages.success(request, "Asignación aceptada correctamente.")

        elif estado == 'Rechazada':
            if not comentario_rechazo:
                messages.error(request, "Debe proporcionar un motivo para rechazar la asignación.")
                return redirect('gestionar_asignaciones')

            RespuestaAsignacion_taller.objects.create(
                asignacion=asignacion,
                usuario=request.user,
                estado=estado,
                comentario_rechazo=comentario_rechazo,
            )
            messages.success(request, "Asignación rechazada correctamente.")

        return redirect('gestionar_asignaciones')

    return render(request, 'taller/gestionar_asignaciones.html', {'asignaciones': asignaciones})

@role_required(['Taller'])
def lista_unidades_aceptadas(request):
    grupo_usuario = request.user.group

    if grupo_usuario:
        unidades_aceptadas = UnidadAceptada.objects.filter(taller=grupo_usuario)
    else:
        unidades_aceptadas = UnidadAceptada.objects.none()

    return render(request, 'taller/lista_unidades.html', {
        'unidades_aceptadas': unidades_aceptadas  # Pasa el diccionario a la plantilla
    })


def subir_a_s3_Unidad(archivo, carpeta="ot_reparacion"):
    """
    Sube un archivo a S3 y retorna la key generada.
    El parámetro 'carpeta' permite organizar los archivos en subcarpetas.
    """
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        # Obtener la extensión del archivo
        extension = archivo.name.split('.')[-1]
        # Generar un nombre único para el archivo
        nombre_archivo = f"{carpeta}/{uuid.uuid4()}.{extension}"
        
        # Subir el archivo a S3
        s3_client.upload_fileobj(archivo, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)
        print(f"Archivo subido correctamente: {nombre_archivo}")
        return nombre_archivo  # Retorna la key del archivo en S3
    except Exception as e:
        print(f"Error al subir archivo a S3: {e}")
        return None

@role_required(['Taller'])
def editar_unidad_aceptada(request, unidad_id):
    """
    Vista para editar una Unidad Aceptada. Se procesa el formulario y si se sube un nuevo archivo
    en el campo `registro`, se almacena en S3 antes de guardar los cambios.
    """
    unidad = get_object_or_404(UnidadAceptada, id=unidad_id)

    # Verificar que el usuario pertenece al taller (grupo) de la unidad
    if request.user.group != unidad.taller:
        return redirect('unauthorized_access')  # Redirigir si no pertenece al grupo

    if request.method == 'POST':
        form = UnidadAceptadaForm(request.POST, request.FILES, instance=unidad)
        if form.is_valid():
            # Obtener el archivo del campo 'registro'
            registro_file = request.FILES.get('registro')
            if registro_file:
                s3_key_registro = subir_a_s3_Unidad(registro_file, carpeta="ot_reparacion")
                if s3_key_registro:
                    unidad.registro = s3_key_registro
                else:
                    messages.error(request, 'Error al subir el documento a S3.')
                    return redirect('editar_unidad_aceptada', unidad_id=unidad.id)

            form.save()
            messages.success(request, 'Unidad aceptada actualizada correctamente.')
            return redirect('lista_unidades_aceptadas')  # Redirigir a la lista de unidades aceptadas
    else:
        form = UnidadAceptadaForm(instance=unidad)

    return render(request, 'taller/editar_unidad.html', {'form': form, 'unidad': unidad})
#Fin Taller Usuario

#Vista Desempeño Empresa
@role_required(['Empresa'])
def homeEmpresa(request):
    # Obtener el grupo (empresa) del usuario logueado
    empresa = request.user.group  # Asumiendo que el usuario está asignado a un grupo
    
    # Obtener todos los hallazgos cerrados de esa empresa
    hallazgos_cerrados = HallazgoEmpresa.objects.filter(grupo=empresa, estado='Cerrado')
    
    # Calcular el desempeño promedio
    total_desempeno = 0
    total_hallazgos = hallazgos_cerrados.count()
    
    if total_hallazgos > 0:
        for hallazgo in hallazgos_cerrados:
            if hallazgo.clasificacion_tiempo_cierre == 'Efectivo':
                total_desempeno += 100
            elif hallazgo.clasificacion_tiempo_cierre == 'Regular':
                total_desempeno += 60
            elif hallazgo.clasificacion_tiempo_cierre == 'Ineficiente':
                total_desempeno += 0
        
        desempeño_promedio = total_desempeno / total_hallazgos
    else:
        desempeño_promedio = 0  # Si no hay hallazgos cerrados, el desempeño es 0

    # Determinar el estado del desempeño
    if desempeño_promedio >= 94:
        estado = 'Excelencia'
        color = 'green'  # Verde
    elif 85 <= desempeño_promedio < 94:
        estado = 'Revisar Procesos'
        color = 'yellow'  # Amarillo
    elif 61 <= desempeño_promedio < 85:
        estado = 'Mala Gestión'
        color = 'red'  # Rojo
    else:
        estado = 'Evaluar Continuidad'
        color = 'darkred'  # Rojo más intenso

    # Pasar los datos a la plantilla
    return render(request, 'empresa/home.html', {
        'desempeño_promedio': desempeño_promedio,
        'estado': estado,
        'color': color,
    })
#Fin Vista Empresa

#Vista Visualizador
def homeVisual(request):
    return render(request, 'visualizador/homevi.html')
#Fin Vista Visualizador

#Vista Taller
def homeTallerUsuario(request):
    return render(request, 'taller/hometaller.html')
#Fin Vista Taller

#Encriptar

# Encriptar un mensaje
def encrypt(user):
    key = Fernet.generate_key()
    print(f"Generated Key: {key.decode()}")
    # Convertir el objeto User a un diccionario
    user_dict = {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "role_name": user.role.name,
        "id_role": user.role.id
        # Agrega otros campos que necesites
    }
    # Convertir el diccionario a una cadena JSON
    user_json = json.dumps(user_dict)
    # Convertir la cadena JSON a bytes
    user_bytes = user_json.encode('utf-8')
    # Obtener la clave secreta desde las configuraciones
    key = config('SECRET_KEY_TOKEN')
    f = Fernet(key)
    # Encriptar los bytes
    encrypted = f.encrypt(user_bytes)
    return encrypted

# Desencriptar un mensaje
def decrypt_message(encrypted_message):
    # Obtener la clave secreta desde las configuraciones
    key = config('SECRET_KEY_TOKEN').encode()  # Convertir la clave a bytes
    f = Fernet(key)
    # Desencriptar el mensaje
    decrypted_message = f.decrypt(encrypted_message).decode('utf-8')
    # Convertir el string desencriptado a un objeto JSON
    decrypted_json = json.loads(decrypted_message)
    return decrypted_json

# Generar y guardar la clave (solo una vez)

#Fin Encriptar

#API
@role_required(['Administrador', 'Visualizador'])
def get_vehiculos_con_seguimiento(request):
    # Obtener todas las placas de los vehículos registrados en el modelo Vehiculo
    placas_vehiculos = Vehiculo.objects.values_list('patente', flat=True)
    
    # Filtrar los vehículos en VehiculoAPI que tengan una placa registrada en Vehiculo
    vehiculos_con_seguimiento = VehiculoAPI.objects.filter(placa__in=placas_vehiculos)
    
    # Pasar los datos al template
    return render(request, 'areas/gps/vehiculos_seguimiento.html', {'vehiculos': vehiculos_con_seguimiento})

@role_required(['Administrador'])
def make_post_request(request):
    def normalize_plate(plate):
        """Normaliza la placa eliminando espacios, convirtiéndola a mayúsculas y eliminando la 'I' final."""
        if plate:
            plate = plate.strip().upper()
            if plate.endswith("I"):
                plate = plate[:-1]
        return plate

    def get_vehicle_status(last_signal_time):
        """Determina si el vehículo está online u offline según la última señal recibida."""
        if last_signal_time:
            last_signal_datetime = parser.parse(last_signal_time)
            now_utc = datetime.now(pytz.utc)
            return 'offline' if (now_utc - last_signal_datetime) > timedelta(days=7) else 'online'
        return 'offline'

    # **1. Obtener los vehículos registrados en `Vehiculo` y normalizar sus placas**
    vehiculos_dict = {normalize_plate(v.patente): v for v in Vehiculo.objects.all()}

    # **2. Obtener los vehículos ya guardados en `VehiculoAPI`**
    existing_vehicles = {normalize_plate(v.placa): v for v in VehiculoAPI.objects.all()}

    url = 'https://www.drivetech.pro/api/v1/get_vehicles_positions/'
    tokengps = config('API_TOKEN')
    headers = {'Authorization': f'Token {tokengps}', 'Content-Type': 'application/json'}
    body = {'client_id': 'coca-cola_andina'}

    response = requests.post(url, headers=headers, json=body)

    if response.status_code == 200:
        data = response.json()
        new_vehicles = []  # Lista de nuevos vehículos a insertar
        updated_vehicles = []  # Lista de vehículos a actualizar

        for position in data.get('positions', []):
            placa = normalize_plate(position.get('plate', 'N/A'))

            # **3. Filtrar solo los vehículos registrados**
            if placa in vehiculos_dict:
                vehiculo = vehiculos_dict[placa]  # Obtener el objeto Vehiculo asociado
                estado = get_vehicle_status(position.get('datetime'))
                latitud = position.get('latitude')
                longitud = position.get('longitude')

                # **4. Actualizar o crear en `VehiculoAPI`**
                if placa in existing_vehicles:
                    existing_vehicle = existing_vehicles[placa]
                    existing_vehicle.latitud = latitud
                    existing_vehicle.longitud = longitud
                    existing_vehicle.fecha_hora = position.get('datetime')
                    existing_vehicle.odometro = position.get('odometer')
                    existing_vehicle.estado = estado
                    updated_vehicles.append(existing_vehicle)
                else:
                    vehiculo_api = VehiculoAPI(
                        vehiculo=vehiculo,  # Relación con el modelo Vehiculo
                        placa=placa,
                        latitud=latitud,
                        longitud=longitud,
                        fecha_hora=position.get('datetime'),
                        odometro=position.get('odometer'),
                        estado=estado,
                    )
                    new_vehicles.append(vehiculo_api)

        # **5. Guardar cambios en la base de datos**
        if new_vehicles:
            VehiculoAPI.objects.bulk_create(new_vehicles)

        if updated_vehicles:
            VehiculoAPI.objects.bulk_update(updated_vehicles, ['latitud', 'longitud', 'fecha_hora', 'odometro', 'estado'])

        return JsonResponse({'mensaje': 'Datos actualizados correctamente', 'vehiculos': [v.placa for v in new_vehicles + updated_vehicles]})

    else:
        return JsonResponse({'error': f'Error al obtener datos: {response.text}'}, status=500)

@role_required(['Administrador', 'Visualizador'])
def listar_estado_gps(request):
    # Filtrar por estado si se pasa un parámetro
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        estados = EstadoGPS.objects.filter(estado=estado_filter)
    else:
        estados = EstadoGPS.objects.all()

    # Paginación
    paginator = Paginator(estados, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Actualización masiva
    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        vehiculos_seleccionados = request.POST.getlist('vehiculos_seleccionados')

        if nuevo_estado and vehiculos_seleccionados:
            EstadoGPS.objects.filter(id__in=vehiculos_seleccionados).update(estado=nuevo_estado)
            messages.success(request, 'Estados actualizados correctamente.')
            return redirect('lista_estado_gps')  # Redirigir para evitar reenvío del formulario

    return render(request, 'areas/gps/lista_estado_gps.html', {'page_obj': page_obj})

#Fin API

#Desempeño

def obtener_desempeno_por_empresa():
    # Obtener datos de desempeño
    resultados = Documento.objects.filter(
        vehiculo__group__isnull=False
    ).values('vehiculo__group__name').annotate(
        efectividad=Count('id', filter=Q(fecha_vencimiento__gte=date.today())),
        regularidad=Count('id', filter=Q(fecha_vencimiento__lt=date.today(), fecha_vencimiento__gte=date.today() - timedelta(days=5))),
        ineficiencia=Count('id', filter=Q(fecha_vencimiento__lt=date.today() - timedelta(days=5))),
    )
    return resultados

def dashboard_view(request):
    desempeno_empresas = obtener_desempeno_por_empresa()
    documentos_vencidos = obtener_documentos_vencidos_por_empresa()
    hallazgos = obtener_hallazgos_por_empresa()
    
    return render(request, 'dashboard/dashboard.html', {
        'desempeno_empresas': desempeno_empresas,
        'documentos_vencidos': documentos_vencidos,
        'hallazgos': hallazgos,
    })

#Fin Desempeño

#EMPRESA
# Vista para cerrar un hallazgo
def subir_a_s3_Cierre(archivo, carpeta="cierre"):
    """
    Sube un archivo a S3 y retorna la key generada.
    El parámetro 'carpeta' permite organizar los archivos en subcarpetas.
    """
    try:
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        # Obtener la extensión del archivo
        extension = archivo.name.split('.')[-1]
        # Generar un nombre único para el archivo
        nombre_archivo = f"{carpeta}/{uuid.uuid4()}.{extension}"
        
        # Subir el archivo a S3
        s3_client.upload_fileobj(archivo, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)
        print(f"Archivo subido correctamente: {nombre_archivo}")
        return nombre_archivo  # Retorna la key del archivo en S3
    except Exception as e:
        print(f"Error al subir archivo a S3: {e}")
        return None

@role_required(['Empresa'])
def cerrar_hallazgo(request, hallazgo_id):
    """
    Vista para cerrar un hallazgo. Se procesa el formulario de cierre,
    se suben los archivos (si existen) a S3 usando la función subir_a_s3,
    se asocia el cierre al hallazgo y se actualiza su estado a 'Cerrado'.
    Luego se redirige a la vista de detalle para visualizar el cierre.
    """
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)

    # Verificar que el hallazgo esté pendiente antes de permitir el cierre
    if hallazgo.estado != 'Pendiente':
        messages.error(request, 'El hallazgo ya está cerrado o no está en estado pendiente.')
        return redirect('listar_hallazgo')

    if request.method == 'POST':
        form = CierreForm(request.POST, request.FILES)
        if form.is_valid():
            # Crear un cierre asociado al hallazgo sin guardar aún
            cierre = form.save(commit=False)
            cierre.hallazgo = hallazgo

            # Subir a S3 la evidencia del cierre, si se envió en el formulario
            evidencia_file = request.FILES.get('evidencia_cierre')
            if evidencia_file:
                s3_key_evidencia = subir_a_s3_Cierre(evidencia_file, carpeta="cierre/evidencia")
                if s3_key_evidencia:
                    cierre.evidencia_cierre = s3_key_evidencia
                else:
                    messages.error(request, 'Error al subir la evidencia a S3.')
                    return redirect('cerrar_hallazgo', hallazgo_id=hallazgo.id)

            # Subir a S3 el documento del cierre, si se envió en el formulario
            documento_file = request.FILES.get('documento_cierre')
            if documento_file:
                s3_key_documento = subir_a_s3_Cierre(documento_file, carpeta="cierre/documento")
                if s3_key_documento:
                    cierre.documento_cierre = s3_key_documento
                else:
                    messages.error(request, 'Error al subir el documento a S3.')
                    return redirect('cerrar_hallazgo', hallazgo_id=hallazgo.id)

            # Guardar el objeto Cierre
            cierre.save()

            # Actualizar el estado del hallazgo a 'Cerrado'
            hallazgo.cerrar_hallazgo()

            # Enviar notificación de cierre por correo
            from .utils import enviar_notificacion_grupo
            enviar_notificacion_grupo(hallazgo, 'Hallazgo Cerrado', 'El hallazgo ha sido cerrado', tipo='cierre')

            messages.success(request, 'Hallazgo cerrado exitosamente.')
            # Redirigir a la vista de detalle para visualizar el cierre y las URL temporales
            return redirect('detalle_hallazgo', hallazgo_id=hallazgo.id)
    else:
        form = CierreForm()

    return render(request, 'empresa/cerrar_hallazgo.html', {'form': form, 'hallazgo': hallazgo})

@role_required(['Empresa'])
def listar_hallazgo(request):
    # Obtener el grupo del usuario actual
    grupo_usuario = request.user.group  # Asumiendo que el campo 'group' está en el modelo User

    # Filtrar los hallazgos según el grupo del usuario
    hallazgos = HallazgoEmpresa.objects.filter(grupo=grupo_usuario)

    return render(request, 'empresa/hallazgos_list.html', {'hallazgos': hallazgos})

"""
@role_required(['Empresa'])
def detalle_hallazgo(request, hallazgo_id):
    # Obtener el hallazgo correspondiente
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)

    # Verificar si existe un cierre relacionado con este hallazgo
    cierre = Cierre.objects.filter(hallazgo=hallazgo).first()  # Devuelve el primer cierre o None

    archivo_s3 = str(hallazgo.evidencia.name)
    print(f"Archivo S3: {archivo_s3}")

    tiempo_expiracion = 240

    s3_client = boto3.client(
        's3',
        region_name=settings.AWS_S3_REGION_NAME,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    # Generar URL temporal
    try:
        url_temporal = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
            ExpiresIn=tiempo_expiracion
        )
        print(f"URL temporal generada: {url_temporal}")   
    except Exception as e:
        print(f"Error al generar el enlace temporal: {e}")
        url_temporal = None

    return render(request, 'empresa/detalle_hallazgo.html', {
        'hallazgo': hallazgo,
        'cierre': cierre,
        'url_temporal': url_temporal
    })

#FIN EMPRESA
"""

@role_required(['Empresa'])
def detalle_hallazgo(request, hallazgo_id):
    hallazgo = get_object_or_404(HallazgoEmpresa, id=hallazgo_id)
    cierre = Cierre.objects.filter(hallazgo=hallazgo).first()  # Devuelve el primer cierre o None

    url_evidencia_hallazgo = None
    url_evidencia = None
    url_documento = None

    if hallazgo.evidencia:
        archivo_s3 = str(hallazgo.evidencia.name)  # Obtener la key del archivo en S3

        tiempo_expiracion = 240  # Tiempo de expiración en segundos

        # Creamos el cliente de S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )

        # Generamos la URL temporal
        try:
            url_evidencia_hallazgo = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
                ExpiresIn=tiempo_expiracion
            )
        except Exception as e:
            print(f"Error al generar URL temporal para evidencia: {e}")
            url_evidencia_hallazgo = None

    if cierre:
        # Crear el cliente de S3
        s3_client = boto3.client(
            's3',
            region_name=settings.AWS_S3_REGION_NAME,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
        )
        tiempo_expiracion = 240  # Tiempo de expiración en segundos

        # Generar URL temporal para evidencia_cierre si existe
        if cierre.evidencia_cierre:
            try:
                url_evidencia = s3_client.generate_presigned_url(
                    'get_object',
                    Params={
                        'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                        'Key': str(cierre.evidencia_cierre.name)
                    },
                    ExpiresIn=tiempo_expiracion
                )
            except Exception as e:
                print(f"Error al generar URL temporal para evidencia_cierre: {e}")
                url_evidencia = None

        # Generar URL temporal para documento_cierre si existe
        if cierre.documento_cierre:
            try:
                url_documento = s3_client.generate_presigned_url(
                    'get_object',
                    Params={
                        'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                        'Key': str(cierre.documento_cierre.name)
                    },
                    ExpiresIn=tiempo_expiracion
                )
            except Exception as e:
                print(f"Error al generar URL temporal para documento_cierre: {e}")
                url_documento = None

    return render(request, 'empresa/detalle_hallazgo.html', {
        'hallazgo': hallazgo,
        'cierre': cierre,
        'url_evidencia': url_evidencia,
        'url_documento': url_documento,
        'url_evidencia_hallazgo': url_evidencia_hallazgo
    })

#FIN EMPRESA

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)

def permission_denied_view(request):
    return render(request, 'error/permission_denied.html')

#Home

def homeACComercial(request):
    return render(request,'ACComercial/homeac.html')


def lista_conductores(request):
    user_group = request.user.group
    conductores = Conductor.objects.filter(empresa=user_group) if user_group else Conductor.objects.none()
    
    today = now().date()
    limite_por_vencer = today + timedelta(days=15)

    # Agregar estado de licencia y color dinámicamente
    for conductor in conductores:
        if conductor.FechaVencimientoLicencia < today:
            conductor.estado_licencia = "Vencida"
            conductor.color_estado = "#e6263c"  # Rojo
        elif today <= conductor.FechaVencimientoLicencia <= limite_por_vencer:
            conductor.estado_licencia = "Por vencer"
            conductor.color_estado = "#E4EB1C"  # Amarillo
        else:
            conductor.estado_licencia = "Vigente"
            conductor.color_estado = "#262d5e"  # Verde

    # Filtrar por estado de la licencia
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado:
        conductores = [c for c in conductores if c.estado_licencia == filtro_estado]

    # Paginación (10 conductores por página)
    paginator = Paginator(conductores, 10)
    page_number = request.GET.get('page')
    conductores_page = paginator.get_page(page_number)

    return render(request, 'empresa/conductores/lista_conductores.html', {
        'conductores': conductores_page,
        'today': today,
        'filtro_estado': filtro_estado
    })

@role_required(['Empresa'])
def editar_licencia_empresa(request, conductor_id):
    # Obtener el conductor que corresponde al grupo del usuario (empresa)
    conductor = get_object_or_404(Conductor, id=conductor_id, empresa=request.user.group)

    # Obtener las licencias actuales del conductor
    licencias = LicenciaConductor.objects.filter(conductor=conductor)

    if request.method == "POST":
        archivos = request.FILES.getlist('archivo_licencia')  # Obtener todos los archivos subidos

        # Validar si se ha subido al menos un archivo
        if archivos:
            # Eliminar las licencias anteriores si no se selecciona mantener
            if not request.POST.get('mantener_licencias'):
                # Imprimir para depurar cuáles licencias se van a eliminar
                licencias_eliminadas = list(licencias.values())
                print("Licencias a eliminar:", licencias_eliminadas)

                # Instanciar el cliente de S3 para poder eliminar los archivos
                s3_client = boto3.client(
                    's3',
                    region_name=settings.AWS_S3_REGION_NAME,
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
                )

                # Recorrer cada licencia y eliminar su archivo en S3
                for licencia in licencias:
                    s3_key = licencia.archivo.name  # La clave (key) en S3
                    try:
                        s3_client.delete_object(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                            Key=s3_key
                        )
                        print(f"Archivo eliminado de S3: {s3_key}")
                    except Exception as e:
                        print(f"Error al eliminar {s3_key} de S3: {e}")

                # Eliminar los registros de la base de datos
                licencias.delete()

            # Subir los nuevos archivos a S3 y guardarlos en la BD
            for archivo in archivos:
                # Utilizamos la función que se encarga de subir a S3 y retornar la URL
                url_archivo = subir_a_s3(archivo, "licencias")
                if url_archivo:
                    LicenciaConductor.objects.create(
                        conductor=conductor,
                        archivo=url_archivo  # Guardamos la URL en el campo correspondiente
                    )
                else:
                    messages.error(request, "Hubo un error al subir uno de los archivos a S3.")
                    return redirect('editar_licencia_empresa', conductor_id=conductor.id)
                    
            messages.success(request, "Licencias actualizadas exitosamente.")
            return redirect('lista_conductores')  # Redirigir a la lista de conductores
        else:
            messages.error(request, "Por favor, sube al menos un archivo de licencia.")


    return render(request, 'empresa/conductores/editar_licencia.html', {
        'conductor': conductor,
        'licencias': licencias  # Pasar las licencias existentes al template
    })


def licencia_detalle_empresa(request, licencia_id):
    # Se asegura que el registro consultado pertenezca a un conductor de la empresa del usuario.
    licencia = get_object_or_404(LicenciaConductor, id=licencia_id, conductor__empresa=request.user.group)

    # Se obtiene el nombre del archivo almacenado en S3.
    archivo_s3 = str(licencia.archivo.name)
    print(f"Archivo S3: {archivo_s3}")

    # Tiempo de expiración para el URL temporal (en segundos)
    tiempo_expiracion = 240

    # Instanciar el cliente S3 con las credenciales y configuración de settings
    s3_client = boto3.client(
        's3',
        region_name=settings.AWS_S3_REGION_NAME,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )

    # Generar un URL temporal para acceder al archivo en S3
    try:
        url_temporal = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': archivo_s3},
            ExpiresIn=tiempo_expiracion
        )
    except Exception as e:
        print(f"Error al generar URL temporal: {e}")
        url_temporal = None

    # Renderiza el template de detalle (por ejemplo, en 'empresa/conductores/licencia_detalle.html')
    return render(request, 'empresa/conductores/licencia_detalle_empresa.html', {
        'licencia': licencia,
        'url_temporal': url_temporal
    })

